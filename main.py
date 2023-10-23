import openpyxl
from openpyxl.styles import Alignment
import asyncio
import logging
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters.command import Command
from aiogram.filters import CommandObject
from datetime import datetime

workbook = openpyxl.load_workbook('example.xlsx') # Открываем файл Excel
sheet = workbook.active# Выбираем активный лист
today = datetime.today().strftime('%d-%m-%Y')

logging.basicConfig(level=logging.INFO) # Включаем логирование, чтобы не пропустить важные сообщения

bot = Bot(token="6778184881:AAHwWFtZ1hlDWljgmj_GwUBra0NjU82gW8c") # Объект бота
dp = Dispatcher() # Диспетчер

#-------------------------------------------Функции--------------------------------

async def add_row(message_text: str): # функция для добавления строки с данными об упражнении
    
    data = message_text.split(',')

    data_to_write = [data[0], data[1], data[2]] # Данные, которые вы хотите записать

    next_row = sheet.max_row + 1 # Находим следующую свободную строку в столбце A
   
    for idx, value in enumerate(data_to_write, start=1): # Записываем данные в следующую свободную строку в столбцы A, B, C и так далее
        sheet.cell(row=next_row, column=idx, value=value)

    workbook.save("example.xlsx") # Сохраняем изменения в файле

    return data[0], data[1], data[2]

#-------------------------------------------Функции--------------------------------


#---------------------------------------Обработка сообщений--------------------------------

@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    kb = [
        [
            types.KeyboardButton(text="Начать новую тренировку"),
            types.KeyboardButton(text="Просмотр предыдущих")
        ],
    ]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )
    await message.answer("привет, я бот для записи твоих тренировок. начнем тренировку?", reply_markup=keyboard)



@dp.message(F.text.lower() == "начать новую тренировку")
async def with_puree(message: types.Message):
    # await message.reply("пришлите название упражнения, количество повторений и вес на подходе в таком формате: жим лежа, 8, 40")

    await message.reply("Отлично! Пришлите название упражнения, количество повторений и вес на подходе в таком формате: жим лежа, 8, 60", reply_markup=types.ReplyKeyboardRemove())
    # Находим следующую свободную строку в столбце A
    next_row = sheet.max_row + 1

    # Объединяем ячейки в строке A от A{next_row} до C{next_row}
    sheet.merge_cells(f'A{next_row}:C{next_row}')

    # Записываем сегодняшнюю дату в объединенную ячейку
    sheet[f'A{next_row}'] = today
    alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'A{next_row}'].alignment = alignment


@dp.message()
async def handle_message(message: types.Message):
    # Сохраняем сообщение в переменную
    message_text = message.text

    add_row(message_text)

    name_ex, count, weight = await add_row(message_text)  # Вызываем функцию add_row и сохраняем возвращаемые значения в переменных

    await message.answer(f'Записал данные об упражнении. Название упражнения: {name_ex}, количество повторений: {count} , вес на подходе: {weight}')



# Запуск процесса поллинга новых апдейтов
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())


